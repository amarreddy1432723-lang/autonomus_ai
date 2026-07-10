import csv
import hashlib
import io
import json
import os
import re
import uuid
from pathlib import Path
from typing import Iterable
from uuid import UUID

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from services.shared.models import FileChunk, FileReference
from .config import settings


SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".json", ".csv", ".py", ".js", ".ts", ".tsx", ".html", ".css",
    ".pdf", ".docx", ".xlsx",
}

PROFILE_SECTION_KEYWORDS = {
    "education": ("education", "academic", "qualification", "degree", "university", "college", "school"),
    "skills": ("skill", "technology", "tools", "framework", "language", "technical"),
    "projects": ("project", "portfolio", "built", "developed", "implemented"),
    "experience": ("experience", "intern", "employment", "work", "role", "responsibility"),
    "achievements": ("achievement", "award", "certification", "metric", "impact", "rank", "score"),
}


def estimate_tokens(text: str) -> int:
    try:
        import tiktoken

        encoding = tiktoken.get_encoding("cl100k_base")
        return len(encoding.encode(text or ""))
    except Exception:
        return max(1, len((text or "").split()))


def _storage_root() -> Path:
    root = Path(settings.FILE_STORAGE_LOCAL_DIR)
    if not root.is_absolute():
        root = Path(__file__).resolve().parents[3] / root
    root.mkdir(parents=True, exist_ok=True)
    return root


def _s3_client():
    import boto3

    return boto3.client(
        "s3",
        endpoint_url=settings.S3_ENDPOINT_URL,
        aws_access_key_id=settings.S3_ACCESS_KEY_ID,
        aws_secret_access_key=settings.S3_SECRET_ACCESS_KEY,
        region_name=settings.S3_REGION,
    )


def storage_provider() -> str:
    if (
        settings.FILE_STORAGE_PROVIDER.lower() == "s3"
        and settings.S3_BUCKET
        and settings.S3_ACCESS_KEY_ID
        and settings.S3_SECRET_ACCESS_KEY
    ):
        return "s3"
    return "local"


def put_object(object_key: str, data: bytes, content_type: str | None) -> None:
    if storage_provider() == "s3":
        _s3_client().put_object(
            Bucket=settings.S3_BUCKET,
            Key=object_key,
            Body=data,
            ContentType=content_type or "application/octet-stream",
        )
        return

    path = _storage_root() / object_key
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def get_object(object_key: str) -> bytes:
    if storage_provider() == "s3":
        response = _s3_client().get_object(Bucket=settings.S3_BUCKET, Key=object_key)
        return response["Body"].read()
    path = _storage_root() / object_key
    if not path.exists():
        raise HTTPException(status_code=404, detail="Stored object not found")
    return path.read_bytes()


def delete_object(object_key: str) -> None:
    if storage_provider() == "s3":
        _s3_client().delete_object(Bucket=settings.S3_BUCKET, Key=object_key)
        return
    path = _storage_root() / object_key
    if path.exists():
        path.unlink()


async def create_file_reference(
    db: Session,
    user_id: UUID,
    upload: UploadFile,
    owner_type: str = "chat",
    owner_id: UUID | None = None,
) -> FileReference:
    filename = Path(upload.filename or "upload").name
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {extension or 'unknown'}")

    data = await upload.read()
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File exceeds 25 MB limit")

    checksum = hashlib.sha256(data).hexdigest()
    object_key = f"users/{user_id}/files/{uuid.uuid4()}{extension}"
    put_object(object_key, data, upload.content_type)

    record = FileReference(
        user_id=user_id,
        owner_type=owner_type,
        owner_id=owner_id,
        storage_provider=storage_provider(),
        bucket=settings.S3_BUCKET,
        object_key=object_key,
        filename=filename,
        content_type=upload.content_type,
        size_bytes=len(data),
        checksum_sha256=checksum,
        status="active",
        metadata_json={"extension": extension},
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def _text_from_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _text_from_docx(data: bytes) -> str:
    from docx import Document

    document = Document(io.BytesIO(data))
    return "\n".join(paragraph.text for paragraph in document.paragraphs)


def _text_from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        rows.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            rows.append("\t".join("" if value is None else str(value) for value in row))
    return "\n".join(rows)


def extract_text_from_file(record: FileReference) -> str:
    data = get_object(record.object_key)
    extension = Path(record.filename).suffix.lower()
    if extension == ".pdf":
        return _text_from_pdf(data)
    if extension == ".docx":
        return _text_from_docx(data)
    if extension == ".xlsx":
        return _text_from_xlsx(data)
    if extension == ".csv":
        decoded = data.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(decoded))
        return "\n".join("\t".join(row) for row in reader)
    if extension == ".json":
        decoded = data.decode("utf-8", errors="ignore")
        try:
            return json.dumps(json.loads(decoded), indent=2)
        except Exception:
            return decoded
    return data.decode("utf-8", errors="ignore")


def chunk_text(text: str, max_tokens: int = 700) -> Iterable[tuple[str, int]]:
    paragraphs = [part.strip() for part in (text or "").splitlines() if part.strip()]
    current: list[str] = []
    current_tokens = 0
    for paragraph in paragraphs:
        tokens = estimate_tokens(paragraph)
        if current and current_tokens + tokens > max_tokens:
            content = "\n".join(current)
            yield content, estimate_tokens(content)
            current = []
            current_tokens = 0
        current.append(paragraph)
        current_tokens += tokens
    if current:
        content = "\n".join(current)
        yield content, estimate_tokens(content)


def _clean_resume_line(line: str) -> str:
    return re.sub(r"\s+", " ", line.replace("\t", " ")).strip(" -•|")


def _unique_limited(lines: list[str], limit: int) -> list[str]:
    seen: set[str] = set()
    selected: list[str] = []
    for raw in lines:
        line = _clean_resume_line(raw)
        if len(line) < 3:
            continue
        key = line.lower()
        if key in seen:
            continue
        seen.add(key)
        selected.append(line[:280])
        if len(selected) >= limit:
            break
    return selected


def build_candidate_profile(text: str, filename: str, token_count: int) -> dict:
    from .llm_router import get_chat_llm
    from langchain_core.messages import HumanMessage
    import json

    llm = get_chat_llm(role="reasoning")
    prompt = (
        "You are an expert HR candidate profile parser. Analyze the following resume text and construct a structured candidate profile.\n"
        "Return the output STRICTLY as a JSON object with the following keys:\n"
        "- candidate_name: (string - candidate's name or title if name is not found)\n"
        "- education: (list of strings summarizing degrees, schools, years)\n"
        "- skills: (list of strings of top technical/soft skills)\n"
        "- projects: (list of strings summarizing best projects with tech stack and metrics/outcomes)\n"
        "- experience: (list of strings summarizing work experience, company, role, duration)\n"
        "- strengths: (list of strings outlining candidate strengths)\n"
        "- missing_details: (list of strings representing details like missing dates, metrics, or credentials that would make the profile stronger)\n\n"
        f"Resume text:\n{text}"
    )

    try:
        response = llm.invoke([HumanMessage(content=prompt)], response_format={"type": "json_object"})
        data = json.loads(response.content)
        return {
            "status": "stored",
            "filename": filename,
            "source_tokens": token_count,
            "candidate_summary": f"Profile of {data.get('candidate_name', 'Candidate')} specializing in {', '.join(data.get('skills', [])[:5])}",
            "education": data.get("education") or [],
            "skills": data.get("skills") or [],
            "projects": data.get("projects") or [],
            "experience": data.get("experience") or [],
            "strengths": data.get("strengths") or [],
            "missing_details": data.get("missing_details") or [],
        }
    except Exception as e:
        print(f"LLM resume profile generation failed: {e}. Falling back to empty profile.")
        return {
            "status": "failed",
            "filename": filename,
            "source_tokens": token_count,
            "candidate_summary": "Failed to parse resume with LLM.",
            "education": [],
            "skills": [],
            "projects": [],
            "experience": [],
            "strengths": [],
            "missing_details": [],
        }


def candidate_profile_to_prompt(profile: dict) -> str:
    if not profile:
        return ""

    def section(title: str, value: object) -> str:
        if isinstance(value, list):
            lines = [str(item).strip() for item in value if str(item).strip()]
            return f"{title}:\n" + "\n".join(f"- {line}" for line in lines) if lines else ""
        text = str(value or "").strip()
        return f"{title}:\n{text}" if text else ""

    blocks = [
        f"STORED CANDIDATE PROFILE: {profile.get('filename', 'uploaded resume')}",
        section("Candidate summary", profile.get("candidate_summary")),
        section("Education", profile.get("education")),
        section("Skills", profile.get("skills")),
        section("Projects", profile.get("projects")),
        section("Experience", profile.get("experience")),
        section("Achievements / metrics", profile.get("achievements")),
        section("Missing details to avoid inventing", profile.get("missing_details")),
    ]
    return "\n\n".join(block for block in blocks if block)


def extract_file_to_chunks(db: Session, user_id: UUID, file_id: UUID) -> dict:
    record = db.query(FileReference).filter(
        FileReference.id == file_id,
        FileReference.user_id == user_id,
        FileReference.status == "active",
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="File not found")

    text = extract_text_from_file(record)
    db.query(FileChunk).filter(FileChunk.file_id == record.id, FileChunk.user_id == user_id).delete()
    chunks = []
    for index, (content, token_count) in enumerate(chunk_text(text)):
        chunks.append(FileChunk(
            file_id=record.id,
            user_id=user_id,
            chunk_index=index,
            content=content,
            token_count=token_count,
            metadata_json={"filename": record.filename},
        ))
    db.add_all(chunks)
    total_tokens = sum(c.token_count or 0 for c in chunks)
    profile = build_candidate_profile(text, record.filename, total_tokens)
    record.metadata_json = {
        **(record.metadata_json or {}),
        "extracted": True,
        "chunk_count": len(chunks),
        "token_count": total_tokens,
        "candidate_profile": profile,
    }
    db.commit()
    return {
        "file_id": str(record.id),
        "chunk_count": len(chunks),
        "token_count": total_tokens,
        "candidate_profile_stored": True,
    }


def file_context_for_prompt(db: Session, user_id: UUID, file_ids: list[str], query: str, limit: int = 8) -> str:
    if not file_ids:
        return ""
    parsed_ids = []
    for file_id in file_ids:
        try:
            parsed_ids.append(UUID(str(file_id)))
        except ValueError:
            continue
    if not parsed_ids:
        return ""

    # 1. Fetch chunks for the files first to check total document size
    chunks = db.query(FileChunk, FileReference.filename).join(
        FileReference, FileReference.id == FileChunk.file_id
    ).filter(
        FileChunk.user_id == user_id,
        FileChunk.file_id.in_(parsed_ids),
    ).all()

    # 2. If the total chunks is small (e.g. <= 20, typical for resumes and short files),
    # return the entire original text content in sequential order to preserve structured bullet points
    if len(chunks) <= 20:
        chunks_sorted = sorted(chunks, key=lambda item: (item[0].file_id, item[0].chunk_index))
        blocks = []
        current_file = None
        for chunk, filename in chunks_sorted:
            if filename != current_file:
                current_file = filename
                blocks.append(f"=== FULL FILE: {filename} ===")
            blocks.append(chunk.content)
        return "\n\n".join(blocks)

    # 3. If file size is large, fall back to candidate profile extraction or term-based search chunks
    records = db.query(FileReference).filter(
        FileReference.user_id == user_id,
        FileReference.id.in_(parsed_ids),
        FileReference.status == "active",
    ).all()
    profile_blocks = []
    for record in records:
        profile = (record.metadata_json or {}).get("candidate_profile") or {}
        prompt_profile = candidate_profile_to_prompt(profile)
        if prompt_profile:
            profile_blocks.append(prompt_profile)

    query_lower = (query or "").lower()
    needs_specific_chunks = any(
        phrase in query_lower
        for phrase in ("exact", "specific", "full resume", "line", "section", "all details", "verbatim")
    )
    if profile_blocks and not needs_specific_chunks:
        return "\n\n---\n\n".join(profile_blocks)

    query_terms = {term.lower() for term in (query or "").split() if len(term) > 2}
    ranked = []
    for chunk, filename in chunks:
        content_lower = chunk.content.lower()
        score = sum(1 for term in query_terms if term in content_lower)
        ranked.append((score, chunk, filename))
    ranked.sort(key=lambda item: (item[0], -(item[1].chunk_index or 0)), reverse=True)

    blocks = []
    blocks.extend(profile_blocks)
    for _, chunk, filename in ranked[:limit]:
        blocks.append(f"FILE: {filename}\nCHUNK: {chunk.chunk_index}\n{chunk.content}")
    return "\n\n---\n\n".join(blocks)


def get_file_text(db: Session, user_id: UUID, file_id: UUID) -> str:
    record = db.query(FileReference).filter(FileReference.id == file_id, FileReference.user_id == user_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="File not found")
    if record.owner_type == "code_workspace" and record.owner_id:
        from services.shared.models import CodeSession
        session = db.query(CodeSession).filter(CodeSession.id == record.owner_id).first()
        if session and session.metadata_json and session.metadata_json.get("local_workspace_path"):
            from pathlib import Path
            local_root = Path(str(session.metadata_json["local_workspace_path"])).expanduser().resolve()
            safe_name = record.filename.replace("\\", "/").lstrip("/")
            local_file = (local_root / safe_name).resolve()
            if not str(local_file).startswith(str(local_root)):
                raise HTTPException(status_code=400, detail="Unsafe local workspace file path")
            if local_file.exists():
                return local_file.read_text(encoding="utf-8", errors="ignore")
    return extract_text_from_file(record)
